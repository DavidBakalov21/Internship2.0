
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
from openpyxl.reader.excel import load_workbook
def transliterate_uk(text):
    mapping = {
        'а': 'a',
        'б': 'b',
        'в': 'v',
        'г': 'h',
        'ґ': 'g',
        'д': 'd',
        'е': 'e',
        'є': 'ie',
        'ж': 'zh',
        'з': 'z',
        'и': 'y',
        'і': 'i',
        'ї': 'i',
        'й': 'i',
        'к': 'k',
        'л': 'l',
        'м': 'm',
        'н': 'n',
        'о': 'o',
        'п': 'p',
        'р': 'r',
        'с': 's',
        'т': 't',
        'у': 'u',
        'ф': 'f',
        'х': 'kh',
        'ц': 'ts',
        'ч': 'ch',
        'ш': 'sh',
        'щ': 'shch',
        'ь': '',
        'ю': 'iu',
        'я': 'ia',
        ' ': ' ',
        'А': 'A',
        'Б': 'B',
        'В': 'V',
        'Г': 'H',
        'Ґ': 'G',
        'Д': 'D',
        'Е': 'E',
        'Є': 'Ie',
        'Ж': 'Zh',
        'З': 'Z',
        'И': 'Y',
        'І': 'I',
        'Ї': 'I',
        'Й': 'I',
        'К': 'K',
        'Л': 'L',
        'М': 'M',
        'Н': 'N',
        'О': 'O',
        'П': 'P',
        'Р': 'R',
        'С': 'S',
        'Т': 'T',
        'У': 'U',
        'Ф': 'F',
        'Х': 'Kh',
        'Ц': 'Ts',
        'Ч': 'Ch',
        'Ш': 'Sh',
        'Щ': 'Shch',
        'Ь': '',
        'Ю': 'Iu',
        'Я': 'Ia',
    }

    transliterated_text = ''
    for char in text:
        if char in mapping:
            transliterated_text += mapping[char]
        else:
            transliterated_text += char

    return transliterated_text


def name_correction(name_file_path, output_file_path):
    data_frameTF = pd.read_excel(name_file_path)
    data_frame = pd.read_excel(output_file_path)
    if 'Прізвище' in data_frameTF.columns and "Ім'я" in data_frameTF.columns and "Ім'я(справжнє)" in data_frame.columns:
        surnamesTF = data_frameTF['Прізвище']
        namesTF = data_frameTF["Ім'я"]
        #names = data_frame["Ім'я(справжнє)"]
    else:
        messagebox.showerror("Error", "No required columns found.")
        #print("No required columns")
        return
    full_names = [f"{transliterate_uk(name)}" for name in data_frame["Ім'я(справжнє)"] if pd.notnull(name)]
    full_namesT = [f"{name} {surname}" for name, surname in zip(namesTF, surnamesTF)]
    full_namesTN = [f"{surname} {name} " for name, surname in zip(namesTF, surnamesTF)]
    translitNamesSurnames=[f"{transliterate_uk(name)} {transliterate_uk(surname)}" for name, surname in zip(namesTF, surnamesTF)]
    translitSurnamesNames=[f"{transliterate_uk(surname)} {transliterate_uk(name)}" for name, surname in zip(namesTF, surnamesTF)]


    full_namesT=full_namesT+full_namesTN+translitNamesSurnames+translitSurnamesNames
    #full_namesT = full_namesT + full_namesTN
    ListForWrite = []

    for i in full_names:
        if i in full_namesT:
            ListForWrite.append(i)
        else:
            ListForWrite.append(LevinsteinList(i, full_namesT))
    output_wb = load_workbook(output_file_path)
    output_ws = output_wb.active
    column_name = "Ім'я(справжнє)"
    target_column_name = None
    for column in output_ws.iter_cols(min_row=1, max_row=1):
        if column[0].value == column_name:
            target_column_name = column[0].column
        if target_column_name != None:
            break
    if target_column_name==None:
        messagebox.showerror("Error", "No required columns found.")
        return
    for y, item in enumerate(ListForWrite, start=2):
        output_ws.cell(row=y, column=target_column_name, value=item)  # Assuming name goes in the second column

    output_wb.save(output_file_path)
    messagebox.showinfo("Success","Success")


def LevinsteinList(value, checker):
    draft = {}
    for VARIABLE in checker:
        matrix = np.zeros((len(value) + 1, len(VARIABLE) + 1))

        for i in range(len(value) + 1):
            matrix[i, 0] = i

        for j in range(len(VARIABLE) + 1):
            matrix[0, j] = j

        for i in range(1, len(value) + 1):
            for j in range(1, len(VARIABLE) + 1):
                v1 = matrix[i - 1, j] + 1
                v2 = matrix[i, j - 1] + 1
                v3 = 0
                v4 = 0
                cost = 0

                if value[i - 1] == VARIABLE[j - 1]:
                    v3 = matrix[i - 1, j - 1]
                    cost = 0
                else:
                    v3 = matrix[i - 1, j - 1] + 1
                    cost = 1

                if i > 1 and j > 1:
                    if VARIABLE[j - 1] == value[i - 2] and VARIABLE[j - 2] == value[i - 1]:
                        v4 = matrix[i - 2, j - 2] + cost
                    else:
                        v4 = 999999999
                else:
                    v4 = 999999999

                matrix[i, j] = min(v1, v2, v3, v4)

        addd = matrix[len(value), len(VARIABLE)]
        draft[VARIABLE] = addd

    Sorted_Word_Points = sorted(draft.items(), key=lambda x: x[1])

    res = str(next(iter(Sorted_Word_Points))).split("'")

    return res[1]

def doAlumni(xls_file_path, output_file_path):
    data_frame = pd.read_excel(xls_file_path, sheet_name='Credits + courses taken')
    output_wb = load_workbook(output_file_path)
    output_ws = output_wb.active
    if 'To be taken total' in data_frame.columns and "total" in data_frame.columns and "Students" in data_frame.columns:
        totalPoint = data_frame.iloc[0]['To be taken total']
        deduct = data_frame[data_frame["total"] < totalPoint * 0.8]['Students']
        Notdeduct = data_frame[data_frame["total"] >= totalPoint * 0.8]['Students']
    else:
        messagebox.showerror("Error", "No required columns found in input table.")
        return
    column_deduct = "Відрахувати"
    column_notDeduct = "Залишити"
    target_column_deduct = None
    target_column_notDeduct = None
    for column in output_ws.iter_cols(min_row=1, max_row=1):
        if column[0].value == column_deduct:
            target_column_deduct = column[0].column
        if column[0].value == column_notDeduct:
            target_column_notDeduct = column[0].column
        if target_column_deduct != None and target_column_notDeduct != None:
            break
    if target_column_deduct==None or target_column_notDeduct==None:
        messagebox.showerror("Error", "No required columns found in output table.")
        return

    for i, Deduct in enumerate(deduct, start=2):
        output_ws.cell(row=i, column=target_column_deduct, value=Deduct)
    for i, notdeduct in enumerate(Notdeduct, start=2):
        output_ws.cell(row=i, column=target_column_notDeduct, value=notdeduct)

    output_wb.save(output_file_path)

    messagebox.showinfo("Success","Success")


def doZoom(xls_file_path, output_file_path):
    data_frame = pd.read_excel(xls_file_path)
    if 'Тривалість' in data_frame.columns and "Ім'я(справжнє)" in data_frame.columns:
        duration = data_frame['Тривалість']
        Name = data_frame["Ім'я(справжнє)"]
    else:
        messagebox.showerror("Error", "No required columns found in input table.")
        return
    output_wb = load_workbook(output_file_path)
    output_ws = output_wb.active
    column_duration = "Відвідуваність"
    column_name = "Ім'я(справжнє)"
    target_column_attendance = None
    target_column_name = None
    for column in output_ws.iter_cols(min_row=1, max_row=1):
        if column[0].value == column_duration:
            target_column_attendance = column[0].column
        if column[0].value == column_name:
            target_column_name = column[0].column
        if target_column_attendance != None and target_column_name != None:
            break
    if target_column_name==None or target_column_attendance==None:
        messagebox.showerror("Error", "No required columns found in output table.")
        return
    data = zip(duration, Name)

    for i, (attendance, name) in enumerate(data, start=2):
        status = "absent"
        if attendance >= 46:
            status = "present"
        output_ws.cell(row=i, column=target_column_attendance, value=status)
        output_ws.cell(row=i, column=target_column_name, value=name)

    output_wb.save(output_file_path)
    messagebox.showinfo("Success","Success")


def doMoodle(xls_file_path, output_file_path):
    data_frame = pd.read_excel(xls_file_path)
    if 'Загальне за курс (Бали)' in data_frame.columns and "Прізвище" in data_frame.columns and "Ім'я" in data_frame.columns:
        Grades = data_frame['Загальне за курс (Бали)']
        Surname = data_frame['Прізвище']
        Name = data_frame["Ім'я"]
    else:
        messagebox.showerror("Error", "No required columns found in input table.")
        return
    output_wb = load_workbook(output_file_path)
    output_ws = output_wb.active
    column_grade = "Бали"
    column_surname = "Прізвище"
    column_name = "Ім'я"
    target_column_grade = None
    target_column_name = None
    target_column_surname = None
    for column in output_ws.iter_cols(min_row=1, max_row=1):
        if column[0].value == column_grade:
            target_column_grade = column[0].column
        if column[0].value == column_surname:
            target_column_surname = column[0].column
        if column[0].value == column_name:
            target_column_name = column[0].column
        if target_column_surname != None and target_column_name != None and target_column_grade != None:
            break
    if target_column_name==None or target_column_grade==None or target_column_surname==None:
        messagebox.showerror("Error", "No required columns found in output table.")
        return
    data = zip(Grades, Surname, Name)
    for i, (grade, surname, name) in enumerate(data, start=2):
        output_ws.cell(row=i, column=target_column_grade, value=grade)
        output_ws.cell(row=i, column=target_column_surname, value=surname)
        output_ws.cell(row=i, column=target_column_name, value=name)

    output_wb.save(output_file_path)
    messagebox.showinfo("Success","Success")


def select_input_file():
    global input_file_path
    input_file_path = filedialog.askopenfilename(title="Select input file",
                                                 filetypes=(("Excel files", "*.xls *.xlsx"), ("All files", "*.*")))
    if input_file_path:
        if not (input_file_path.endswith(".xls") or input_file_path.endswith(".xlsx")):
            messagebox.showerror("Error", "Input file must be a .xls or .xlsx file!")
            input_file_path = ""
        else:
            input_file_label['text'] = input_file_path


def select_output_file():
    global output_file_path

    output_file_path = filedialog.askopenfilename(title="Select input file",
                                                  filetypes=(("Excel files", "*.xls *.xlsx"), ("All files", "*.*")))
    if output_file_path:
        if not (output_file_path.endswith(".xls") or output_file_path.endswith(".xlsx")):
            messagebox.showerror("Error", "Input file must be a .xls or .xlsx file!")
            output_file_path = ""
        else:
            output_file_label['text'] = output_file_path


def select_names_file():
    global names_file_path

    names_file_path = filedialog.askopenfilename(title="Select input file",
                                                 filetypes=(("Excel files", "*.xls *.xlsx"), ("All files", "*.*")))
    if names_file_path:
        if not (names_file_path.endswith(".xls") or names_file_path.endswith(".xlsx")):
            messagebox.showerror("Error", "Input file must be a .xls or .xlsx file!")
            names_file_path = ""
        else:
            names_file_label['text'] = names_file_path


def process_files():
    if input_file_path and output_file_path:
        if "Zoom" in input_file_path:
            doZoom(input_file_path, output_file_path)
            if "names_file_path" in globals():
                if names_file_path!="":
                    name_correction(names_file_path, output_file_path)
        elif "Moodle" in input_file_path:
            doMoodle(input_file_path, output_file_path)
        #  name_correction(names_file_path, output_file_path)
        elif "Alumni and Students" in input_file_path:
            doAlumni(input_file_path, output_file_path)

        #print("g")


def open_text_window():
    text_window = tk.Toplevel(root)
    text_window.title("Інструкція")

    text_label = tk.Label(text_window, justify="center", text="""Інструкція користування програмою "KSE Automatization" """,
                          font=('Verdana', 12,'bold'))
    text_label.pack(padx=10, pady=10)
    secondText= tk.Label(text_window, justify="left", text=""" 
    1.Клацніть на кнопку "Select Input File". 
    Відкриється діалогове вікно, в якому ви зможете обрати вхідний файл у форматі .xls або .xlsx. 

    2.Оберіть потрібний файл і натисніть "Open". Над кнопкою "Select Input File" з'явиться шлях до вибраного вхідного файлу. 
    Клацніть на кнопку "Select Output File". Відкриється діалогове вікно, в якому ви зможете обрати вихідний файл у форматі .xls або .xlsx. 

    3.Оберіть потрібний файл і натисніть "Open".
    Над кнопкою "Select Output File" з'явиться шлях до вибраного вихідного файлу. 
    
    4.Натисніть кнопку "Process Files". Програма почне обробку даних. 
    Якщо в назві вхідного файлу присутнє "Zoom", буде запущена функція doZoom(в вхідному файлі необхідні колонки: Тривалість та 
    Ім'я(справжнє), в вихідному: Відвідуваність та Ім'я(справжнє)). 
    Якщо в назві вхідного файлу присутнє "Moodle", буде запущена функція doMoodle(в вхідному файлі необхідні колонки: Загальне 
    за курс (Бали),  Прізвище, Ім'я,  в вихідному: Бали,  Прізвище, Ім'я). 
    Якщо в назві вхідного файлу присутнє "Alumni and Students", буде 
    запущена функція doAlumni(в вхідному файлі необхідні колонки: To be taken total, total, Students  в вихідному: Відрахувати, Залишити). 

    В програмі присутня функція корекції Імен для таблиць "Zoom", для того, щоб її активувати необхідно натиснути 
    на кнопку "Names", після натискання виберіть файл .xls або .xlsx, який містить дві колонки "Прізвище" і "Ім'я" 
    студентів(якщо не обирати файл при натисканні цієї кнопки, то корекція імен не відбудеться).  
    Після обробки файлів в консолі з'явиться повідомлення "Success". Це означає, що дані були успішно оброблені і збережені в вихідному файлі.""",
                          font=('Verdana', 12))
    secondText.pack(padx=10, pady=10)




root = tk.Tk()
root.title("KSE Automatization")

title_label = tk.Label(root, text="KSE Automatization", font=('Verdana', 18, 'bold'))
title_label.pack(pady=10)

explanation_label = tk.Label(root, text="This program helps you systematize tables.", font=('Verdana', 12))
explanation_label.pack(pady=10)

input_file_button = tk.Button(root, text="Select Input File", command=select_input_file, font=('Verdana', 14), padx=20,
                              pady=10)
input_file_button.pack()

input_file_label = tk.Label(root, text="", font=('Verdana', 10))
input_file_label.pack(pady=10)

output_file_button = tk.Button(root, text="Select Output File", command=select_output_file, font=('Verdana', 14),
                               padx=20, pady=10)
output_file_button.pack()

output_file_label = tk.Label(root, text="", font=('Verdana', 10))
output_file_label.pack(pady=10)

process_button = tk.Button(root, text="Process Files", command=process_files, font=('Verdana', 14), padx=20, pady=10)
process_button.pack()
output_label = tk.Label(root, text="", font=('Verdana', 10))
output_label.pack(pady=10)

names_button = tk.Button(root, text="Names", command=select_names_file, font=('Verdana', 14), padx=20, pady=10)
names_button.pack()

names_file_label = tk.Label(root, text="", font=('Verdana', 10))
names_file_label.pack(pady=10)
text_button = tk.Button(root, text="...", command=open_text_window, font=('Verdana', 14), padx=10, pady=5)
text_button.pack()
root.mainloop()
