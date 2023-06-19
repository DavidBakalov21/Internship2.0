import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook

data_frame = pd.read_excel("C:\\Users\Давід\\PycharmProjects\internship\\Students.xlsx")
data_frameT = pd.read_excel("C:\\Users\Давід\\PycharmProjects\internship\\testNames.xlsx")
output_file_path = "C:\\Users\\Давід\\PycharmProjects\\internship\\write.xlsx"
surnames = data_frame['Прізвище']
names = data_frame["Ім'я"]
surnamesT = data_frameT['Прізвище']
namesT = data_frameT["Ім'я"]
full_names = [f"{name} {surname}" for name, surname in zip(names, surnames)]
full_namesT = [f"{name} {surname}" for name, surname in zip(namesT, surnamesT)]

ListForWrite=[]

def LevinsteinList(value, checker):
    draft = {}
    for VARIABLE in checker:
        matrix = np.zeros((len(value) + 1, len(VARIABLE) + 1))

        for i in range(len(value) + 1):
            matrix[i, 0] = i

        for j in range(len(VARIABLE) + 1):
            matrix[0, j] = j

        for i in range(1, len(value) + 1):
            for j in range(1, len(VARIABLE) + 1):
                v1 = matrix[i - 1, j] + 1
                v2 = matrix[i, j - 1] + 1
                v3 = 0
                v4 = 0
                cost = 0

                if value[i - 1] == VARIABLE[j - 1]:
                    v3 = matrix[i - 1, j - 1]
                    cost = 0
                else:
                    v3 = matrix[i - 1, j - 1] + 1
                    cost = 1

                if i > 1 and j > 1:
                    if VARIABLE[j - 1] == value[i - 2] and VARIABLE[j - 2] == value[i - 1]:
                        v4 = matrix[i - 2, j - 2] + cost
                    else:
                        v4 = 999999999
                else:
                    v4 = 999999999

                matrix[i, j] = min(v1, v2, v3, v4)

        addd = matrix[len(value), len(VARIABLE)]
        draft[VARIABLE] = addd

    Sorted_Word_Points = sorted(draft.items(), key=lambda x: x[1])

    res=str(next(iter(Sorted_Word_Points))).split("'")

    return res[1]


for i in full_namesT:
    if i in full_names:
        ListForWrite.append(i)
    else:
        ListForWrite.append(LevinsteinList(i, full_names))

output_wb = load_workbook(output_file_path)
output_ws = output_wb.active
column_surname = "Прізвище"
column_name = "Ім'я"
target_column_name = None
target_column_surname = None
for column in output_ws.iter_cols(min_row=1, max_row=1):
    if column[0].value == column_surname:
        target_column_surname = column[0].column
    if column[0].value == column_name:
        target_column_name = column[0].column
    if target_column_surname != None and target_column_name != None:
        break
for i, item in enumerate(ListForWrite, start=2):
    splited= item.split(" ")
    surname=splited[1]
    name=splited[0]
    output_ws.cell(row=i, column=target_column_surname, value=surname)  # Assuming surname goes in the first column
    output_ws.cell(row=i, column=target_column_name, value=name)  # Assuming name goes in the second column

output_wb.save(output_file_path)
